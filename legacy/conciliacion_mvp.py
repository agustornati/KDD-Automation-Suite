# -*- coding: utf-8 -*-
r"""
conciliacion_mvp.py  -  MVP Conciliacion Bancaria Banco Credicoop (AGENTEX SRL)
================================================================================
Consolida el motor (matching) y el generador de resumen en UN solo script
parametrizado por periodo. No hay nada hardcodeado del mes: el marcador de corte
del extracto, los titulos, la carpeta de salida y el saldo bancario de cierre se
derivan del periodo / se autodetectan del PDF.

USO
---
    python conciliacion_mvp.py --periodo 2025-07 --saldo-contable 21667987.68

    # rutas explicitas (opcional; si no, se autodetectan en la carpeta del periodo)
    python conciliacion_mvp.py --periodo 2025-07 --saldo-contable 21667987.68 \
        --conta "ruta/Credi 07-25.xlsx" --extracto "ruta/Credi 07-25.pdf"

    # sin --saldo-contable: genera los 2 Excel de matching pero NO el Resumen de saldos

Genera en  <base>\MM-YYYY\ :
    1. Conciliacion_Credicoop_MM-YYYY.xlsx     (verde=conciliado / rojo=no, + Resumen)
    2. Conciliacion_NoCoincidentes_MM-YYYY.xlsx (pendientes + gastos a registrar + anulados)
    3. Resumen_Conciliacion_MM-YYYY.xlsx        (conciliacion de saldos, 5 hojas)  [si hay saldo contable]

REGLA contable: Contabilidad DEBE (entrada) <-> Banco CREDITO ;
                Contabilidad HABER (salida)  <-> Banco DEBITO   (lado invertido).
"""

import os
import re
import sys
import glob
import argparse
import calendar
from datetime import date, datetime
from itertools import combinations
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pdfplumber

# ---------------------------------------------------------------------------
BASE_DIR_DEFAULT = r"C:\Users\I550035\Desktop\Mariana"
MESES_ES = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
            "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

AMT = re.compile(r"^\d{1,3}(?:\.\d{3})*,\d{2}$")
DATE = re.compile(r"^\d{2}/\d{2}/\d{2}$")

VERDE = PatternFill("solid", fgColor="C6EFCE")
ROJO = PatternFill("solid", fgColor="FFC7CE")
GRIS = PatternFill("solid", fgColor="F2F2F2")
HDR = PatternFill("solid", fgColor="1A1F5E")
HDRF = Font(bold=True, color="FFFFFF")
AMAR = PatternFill("solid", fgColor="FFE600")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)
MONEY = "#,##0.00"

_CARGOS = ("Impuesto Ley", "SIRCREB", "Recaudacion I.B", "I.V.A.",
           "Percepcion", "Percep.", "Comision", "Com. Exterior",
           "Com. mantenimiento", "GUV", "Comis acred", "Com Acred",
           "Comis acred", "Comision E-CHEQ", "Custodia de Titulos")


# ===========================================================================
# Configuracion derivada del periodo
# ===========================================================================
class Config:
    def __init__(self, periodo, saldo_contable, base_dir,
                 conta=None, extracto=None, saldo_banco=None):
        try:
            y, m = periodo.split("-")
            self.year, self.mm = int(y), int(m)
        except Exception:
            raise SystemExit("ERROR: --periodo debe ser YYYY-MM (ej. 2025-07)")
        self.periodo = periodo
        self.base_dir = base_dir
        self.mmyyyy = "%02d-%04d" % (self.mm, self.year)
        self.month_name = MESES_ES[self.mm]
        self.last_day = calendar.monthrange(self.year, self.mm)[1]
        # marcador de corte del extracto (ej. "SALDO AL 31/07")
        self.stop_marker = "SALDO AL %02d/%02d" % (self.last_day, self.mm)
        self.saldo_label = "al %02d/%02d/%04d" % (self.last_day, self.mm, self.year)

        self.out_dir = os.path.join(base_dir, self.mmyyyy)
        self.out_conc = os.path.join(self.out_dir, "Conciliacion_Credicoop_%s.xlsx" % self.mmyyyy)
        self.out_pend = os.path.join(self.out_dir, "Conciliacion_NoCoincidentes_%s.xlsx" % self.mmyyyy)
        self.out_resumen = os.path.join(self.out_dir, "Resumen_Conciliacion_%s.xlsx" % self.mmyyyy)

        self.conta_xlsx = conta or self._auto_conta()
        self.extracto_pdf = extracto or self._auto_extracto()
        self.saldo_contable = saldo_contable
        self.saldo_banco = saldo_banco   # se resuelve luego (autodetect del PDF)

    def _search_dirs(self):
        # busca en la carpeta del periodo y, si no, en la base
        return [self.out_dir, self.base_dir]

    def _auto_extracto(self):
        for d in self._search_dirs():
            hits = sorted(glob.glob(os.path.join(d, "*.pdf")))
            if hits:
                return hits[0]
        raise SystemExit("ERROR: no encontre PDF de extracto. Pasa --extracto")

    def _auto_conta(self):
        for d in self._search_dirs():
            hits = [f for f in sorted(glob.glob(os.path.join(d, "*.xlsx")))
                    if not os.path.basename(f).startswith(("Conciliacion", "Resumen", "~$"))]
            if hits:
                return hits[0]
        raise SystemExit("ERROR: no encontre Excel de contabilidad. Pasa --conta")


def parse_importe(s):
    return float(s.replace(".", "").replace(",", "."))


def categoria_banco(desc):
    d = desc or ""
    if any(k.lower() in d.lower() for k in _CARGOS):
        return "Gasto/Impuesto bancario"
    return "Operacion"


# ===========================================================================
# 1. Parseo del EXTRACTO (PDF) por coordenadas
# ===========================================================================
def parse_extracto(path, stop_marker):
    movs = []
    with pdfplumber.open(path) as pdf:
        cur_date = None
        stop = False
        for page in pdf.pages:
            if stop:
                break
            words = page.extract_words(use_text_flow=False)
            lines = {}
            for w in words:
                key = round(w["top"] / 3.0)
                lines.setdefault(key, []).append(w)
            for key in sorted(lines):
                toks = sorted(lines[key], key=lambda w: w["x0"])
                texts = [t["text"] for t in toks]
                linea = " ".join(texts)
                if stop_marker in linea:
                    stop = True
                    break
                d = None
                if texts and DATE.match(texts[0]):
                    dd, mm, yy = texts[0].split("/")
                    d = date(2000 + int(yy), int(mm), int(dd))
                    cur_date = d
                deb = cred = None
                for t in toks:
                    if AMT.match(t["text"]):
                        x1 = t["x1"]
                        if x1 < 450:
                            deb = parse_importe(t["text"])
                        elif x1 < 555:
                            cred = parse_importe(t["text"])
                if d is None and deb is None and cred is None:
                    continue
                if deb is None and cred is None:
                    continue
                desc_toks = [t["text"] for t in toks
                             if not AMT.match(t["text"]) and not DATE.match(t["text"])]
                combte = ""
                if desc_toks and re.fullmatch(r"\d{3,}", desc_toks[0]):
                    combte = desc_toks[0]
                    desc_toks = desc_toks[1:]
                desc = " ".join(desc_toks).strip()
                if deb is not None:
                    movs.append({"fecha": cur_date, "combte": combte,
                                 "desc": desc, "tipo": "D", "importe": deb})
                if cred is not None:
                    movs.append({"fecha": cur_date, "combte": combte,
                                 "desc": desc, "tipo": "C", "importe": cred})
    return movs


def saldo_final_pdf(path, stop_marker):
    """Ultimo importe de la linea 'SALDO AL dd/mm' = saldo bancario de cierre."""
    saldo = None
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(use_text_flow=False)
            lines = {}
            for w in words:
                lines.setdefault(round(w["top"] / 3.0), []).append(w)
            for key in sorted(lines):
                toks = sorted(lines[key], key=lambda w: w["x0"])
                linea = " ".join(t["text"] for t in toks)
                amts = [parse_importe(t["text"]) for t in toks if AMT.match(t["text"])]
                if stop_marker in linea and amts:
                    saldo = amts[-1]
    return saldo


# ===========================================================================
# 2. Parseo de la CONTABILIDAD (Excel)
# ===========================================================================
def parse_contabilidad(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if r[0] is None:
            continue
        fec = r[0].date() if isinstance(r[0], datetime) else r[0]
        debe = float(r[5]) if r[5] else 0.0
        haber = float(r[6]) if r[6] else 0.0
        rows.append({
            "fecha": fec, "clase": r[3], "numero": r[4],
            "debe": debe, "haber": haber,
            "cuit": r[7], "sujeto": r[8], "desc": r[9], "subdiario": r[10],
            "tipo_bco": "C" if debe > 0 else "D",
            "importe": debe if debe > 0 else haber,
        })
    return rows


# ===========================================================================
# 2b. Anular autocancelatorios (mismo importe, lados opuestos, misma fecha)
# ===========================================================================
def marcar_anulados(records, side_of):
    by = defaultdict(lambda: {"IN": [], "OUT": []})
    for k, r in enumerate(records):
        r["anulado"] = False
        by[(round(r["importe"], 2), r["fecha"])][side_of(r)].append(k)
    pares = []
    for key, d in by.items():
        for a, b in zip(d["OUT"], d["IN"]):
            records[a]["anulado"] = records[b]["anulado"] = True
            pares.append((a, b))
    return pares


# ===========================================================================
# 3. Matching uno-a-uno (importe + lado invertido, fecha como desempate)
# ===========================================================================
def conciliar(conta, banco):
    idx = {}
    for j, b in enumerate(banco):
        if b.get("anulado"):
            continue
        idx.setdefault((b["tipo"], round(b["importe"], 2)), []).append(j)
    banco_match = [False] * len(banco)
    conta_match = [False] * len(conta)
    pares = []
    for i in sorted(range(len(conta)), key=lambda k: conta[k]["fecha"]):
        c = conta[i]
        if c.get("anulado"):
            continue
        cands = idx.get((c["tipo_bco"], round(c["importe"], 2)), [])
        libres = [j for j in cands if not banco_match[j]]
        if not libres:
            continue
        libres.sort(key=lambda j: abs((banco[j]["fecha"] - c["fecha"]).days)
                    if banco[j]["fecha"] and c["fecha"] else 999)
        j = libres[0]
        banco_match[j] = True
        conta_match[i] = True
        pares.append((i, j))
    return conta_match, banco_match, pares


# ===========================================================================
# 3b. Matching AGRUPADO (N:1) por suma
# ===========================================================================
def _best_group(pool, target, banco, max_k=8, max_span=3):
    n = len(pool)
    if n < 2:
        return None
    best = None
    for k in range(2, min(max_k, n) + 1):
        for comb in combinations(range(n), k):
            if sum(pool[t][1] for t in comb) != target:
                continue
            fechas = [banco[pool[t][0]]["fecha"] for t in comb]
            span = (max(fechas) - min(fechas)).days
            if span > max_span:
                continue
            cand = (span, k, [pool[t][0] for t in comb])
            if best is None or (span, k) < (best[0], best[1]):
                best = cand
        if best is not None and best[0] == 0:
            break
    return best[2] if best else None


def conciliar_grupos(conta, banco, conta_match, banco_match, win=4, cap=22):
    cents = lambda x: int(round(x * 100))
    grupos = {}
    for i, c in enumerate(conta):
        if conta_match[i] or c.get("anulado"):
            continue
        side, tgt = c["tipo_bco"], cents(c["importe"])
        pool = [(j, cents(banco[j]["importe"])) for j in range(len(banco))
                if not banco_match[j] and not banco[j].get("anulado")
                and banco[j]["tipo"] == side
                and categoria_banco(banco[j]["desc"]) == "Operacion"
                and banco[j]["fecha"]
                and abs((banco[j]["fecha"] - c["fecha"]).days) <= win]
        pool.sort(key=lambda p: abs((banco[p[0]]["fecha"] - c["fecha"]).days))
        pool = pool[:cap]
        sel = _best_group(pool, tgt, banco)
        if sel:
            grupos[i] = sel
            for j in sel:
                banco_match[j] = True
            conta_match[i] = "GRUPO"
    return grupos


# ===========================================================================
# 4. Utilidades de estilo
# ===========================================================================
def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR
        cell.font = HDRF
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autofit(ws):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 55)


# ===========================================================================
# 5. Salida 1: Conciliacion (verde/rojo) + Resumen
# ===========================================================================
def generar_excel(conta, banco, conta_match, banco_match, cfg):
    os.makedirs(cfg.out_dir, exist_ok=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Contabilidad"
    cols = ["Fecha", "Clase", "Numero", "Debe", "Haber", "Sujeto",
            "Lado banco", "Estado"]
    ws.append(cols)
    _style_header(ws, len(cols))
    for i, c in enumerate(conta):
        if c.get("anulado"):
            continue
        estado = "CONCILIADO" if conta_match[i] else "NO CONCILIADO"
        ws.append([c["fecha"], c["clase"], c["numero"], c["debe"] or None,
                   c["haber"] or None, c["sujeto"],
                   "Credito" if c["tipo_bco"] == "C" else "Debito", estado])
        fill = VERDE if conta_match[i] else ROJO
        for cc in range(1, len(cols) + 1):
            ws.cell(ws.max_row, cc).fill = fill
            ws.cell(ws.max_row, cc).border = THIN
        ws.cell(ws.max_row, 1).number_format = "DD/MM/YYYY"
        for cc in (4, 5):
            ws.cell(ws.max_row, cc).number_format = MONEY
    _autofit(ws)

    ws2 = wb.create_sheet("Extracto Banco")
    cols2 = ["Fecha", "Combte", "Descripcion", "Debito", "Credito",
             "Estado", "Categoria"]
    ws2.append(cols2)
    _style_header(ws2, len(cols2))
    for j, b in enumerate(banco):
        if b.get("anulado"):
            continue
        estado = "CONCILIADO" if banco_match[j] else "NO CONCILIADO"
        ws2.append([b["fecha"], b["combte"], b["desc"],
                    b["importe"] if b["tipo"] == "D" else None,
                    b["importe"] if b["tipo"] == "C" else None, estado,
                    categoria_banco(b["desc"])])
        fill = VERDE if banco_match[j] else ROJO
        for cc in range(1, len(cols2) + 1):
            ws2.cell(ws2.max_row, cc).fill = fill
            ws2.cell(ws2.max_row, cc).border = THIN
        ws2.cell(ws2.max_row, 1).number_format = "DD/MM/YYYY"
        for cc in (4, 5):
            ws2.cell(ws2.max_row, cc).number_format = MONEY
    _autofit(ws2)

    ws3 = wb.create_sheet("Resumen")
    nc_ok = sum(1 for x in conta_match if x); nc_tot = len(conta)
    nb_ok = sum(1 for x in banco_match if x); nb_tot = len(banco)
    debe = sum(c["debe"] for c in conta); haber = sum(c["haber"] for c in conta)
    bcred = sum(b["importe"] for b in banco if b["tipo"] == "C")
    bdeb = sum(b["importe"] for b in banco if b["tipo"] == "D")
    conc_imp = sum(conta[i]["importe"] for i in range(nc_tot) if conta_match[i])
    nb_nc_oper = sum(1 for j in range(nb_tot)
                     if not banco_match[j] and categoria_banco(banco[j]["desc"]) == "Operacion")
    nb_nc_carg = (nb_tot - nb_ok) - nb_nc_oper
    resumen = [
        ["CONCILIACION BANCARIA - BANCO CREDICOOP - %02d/%04d" % (cfg.mm, cfg.year), ""],
        ["", ""],
        ["CONTABILIDAD", ""],
        ["  Movimientos", nc_tot],
        ["  Conciliados", nc_ok],
        ["  No conciliados", nc_tot - nc_ok],
        ["  Total Debe", debe],
        ["  Total Haber", haber],
        ["", ""],
        ["EXTRACTO BANCO", ""],
        ["  Movimientos", nb_tot],
        ["  Conciliados", nb_ok],
        ["  No conciliados", nb_tot - nb_ok],
        ["    de los cuales operaciones a revisar", nb_nc_oper],
        ["    de los cuales gastos/impuestos bancarios", nb_nc_carg],
        ["  Total Debito", bdeb],
        ["  Total Credito", bcred],
        ["", ""],
        ["Importe conciliado (coincidente)", conc_imp],
        ["Regla", "Conta DEBE <-> Banco CREDITO  |  Conta HABER <-> Banco DEBITO"],
    ]
    for row in resumen:
        ws3.append(row)
    ws3.cell(1, 1).font = Font(bold=True, size=13, color="1A1F5E")
    for r in (3, 10):
        ws3.cell(r, 1).font = Font(bold=True, color="2D3277")
    for r in range(1, ws3.max_row + 1):
        v = ws3.cell(r, 2).value
        if isinstance(v, float):
            ws3.cell(r, 2).number_format = MONEY
    ws3.column_dimensions["A"].width = 42
    ws3.column_dimensions["B"].width = 28

    wb.save(cfg.out_conc)


# ===========================================================================
# 6. Salida 2: No coincidentes (pendientes + gastos a registrar + anulados)
# ===========================================================================
def generar_excel_pendientes(conta, banco, conta_match, banco_match, grupos,
                             cfg, conta_anul=(), banco_anul=()):
    os.makedirs(cfg.out_dir, exist_ok=True)
    wb = openpyxl.Workbook()
    etiqueta = {i: "%s %s" % (conta[i]["clase"], conta[i]["numero"]) for i in grupos}
    banco_a_conta = {}
    for i, js in grupos.items():
        for j in js:
            banco_a_conta[j] = etiqueta[i]

    ws = wb.active
    ws.title = "Contabilidad no 1a1"
    cols = ["Fecha", "Clase", "Numero", "Debe", "Haber", "Sujeto",
            "Lado banco", "Estado", "Detalle grupo (banco)"]
    ws.append(cols)
    _style_header(ws, len(cols))
    for i, c in enumerate(conta):
        if c.get("anulado"):
            continue
        if conta_match[i] is True:
            continue
        if conta_match[i] == "GRUPO":
            estado = "CONCILIADO POR SUMA"
            det = "  +  ".join(
                "{} ${:,.2f}".format(banco[j]["fecha"].strftime("%d/%m"),
                                     banco[j]["importe"])
                for j in grupos[i])
            fill = VERDE
        else:
            estado, det, fill = "NO CONCILIADO", "", ROJO
        ws.append([c["fecha"], c["clase"], c["numero"], c["debe"] or None,
                   c["haber"] or None, c["sujeto"],
                   "Credito" if c["tipo_bco"] == "C" else "Debito", estado, det])
        for cc in range(1, len(cols) + 1):
            ws.cell(ws.max_row, cc).fill = fill
            ws.cell(ws.max_row, cc).border = THIN
        ws.cell(ws.max_row, 1).number_format = "DD/MM/YYYY"
        for cc in (4, 5):
            ws.cell(ws.max_row, cc).number_format = MONEY
    _autofit(ws)

    ws2 = wb.create_sheet("Extracto no 1a1")
    cols2 = ["Fecha", "Combte", "Descripcion", "Debito", "Credito",
             "Categoria", "Estado", "Grupo (conta)"]
    ws2.append(cols2)
    _style_header(ws2, len(cols2))
    for j, b in enumerate(banco):
        if b.get("anulado"):
            continue
        if banco_match[j] is True and j not in banco_a_conta:
            continue
        if j not in banco_a_conta and categoria_banco(b["desc"]) == "Gasto/Impuesto bancario":
            continue
        if j in banco_a_conta:
            estado, grp, fill = "CONCILIADO POR SUMA", banco_a_conta[j], VERDE
        else:
            estado, grp, fill = "NO CONCILIADO", "", ROJO
        ws2.append([b["fecha"], b["combte"], b["desc"],
                    b["importe"] if b["tipo"] == "D" else None,
                    b["importe"] if b["tipo"] == "C" else None,
                    categoria_banco(b["desc"]), estado, grp])
        for cc in range(1, len(cols2) + 1):
            ws2.cell(ws2.max_row, cc).fill = fill
            ws2.cell(ws2.max_row, cc).border = THIN
        ws2.cell(ws2.max_row, 1).number_format = "DD/MM/YYYY"
        for cc in (4, 5):
            ws2.cell(ws2.max_row, cc).number_format = MONEY
    _autofit(ws2)

    ws3 = wb.create_sheet("Gastos a registrar")
    resumen = defaultdict(lambda: [0, 0.0])
    for j, b in enumerate(banco):
        if banco_match[j] is True and j not in banco_a_conta:
            continue
        if j in banco_a_conta:
            continue
        if categoria_banco(b["desc"]) != "Gasto/Impuesto bancario":
            continue
        resumen[b["desc"]][0] += 1
        resumen[b["desc"]][1] += b["importe"]
    cols3 = ["Concepto (descripcion)", "Cant. movimientos", "Importe total"]
    ws3.append(cols3)
    _style_header(ws3, len(cols3))
    total = 0.0
    for desc, (cant, imp) in sorted(resumen.items(), key=lambda x: -x[1][1]):
        ws3.append([desc, cant, imp])
        total += imp
        for cc in range(1, len(cols3) + 1):
            ws3.cell(ws3.max_row, cc).border = THIN
        ws3.cell(ws3.max_row, 3).number_format = MONEY
    ws3.append(["TOTAL", sum(c for c, _ in resumen.values()), total])
    for cc in range(1, len(cols3) + 1):
        ws3.cell(ws3.max_row, cc).font = Font(bold=True)
        ws3.cell(ws3.max_row, cc).fill = AMAR
    ws3.cell(ws3.max_row, 3).number_format = MONEY
    _autofit(ws3)

    ws4 = wb.create_sheet("Anulados (salio y volvio)")
    cols4 = ["Fuente", "Fecha", "Detalle", "Importe", "Contrapartida"]
    ws4.append(cols4)
    _style_header(ws4, len(cols4))
    for a, b in conta_anul:
        ws4.append(["Contabilidad", conta[a]["fecha"],
                    "%s %s (Haber) / %s %s (Debe)" % (
                        conta[a]["clase"], conta[a]["numero"],
                        conta[b]["clase"], conta[b]["numero"]),
                    conta[a]["importe"], "Netea contra si mismo"])
    for a, b in banco_anul:
        ws4.append(["Extracto banco", banco[a]["fecha"],
                    "%s (Debito) / %s (Credito)" % (banco[a]["desc"],
                                                    banco[b]["desc"]),
                    banco[a]["importe"], "Cheque depositado y rechazado"])
    for r in range(2, ws4.max_row + 1):
        ws4.cell(r, 4).number_format = MONEY
        for cc in range(1, len(cols4) + 1):
            ws4.cell(r, cc).border = THIN
        ws4.cell(r, 2).number_format = "DD/MM/YYYY"
    _autofit(ws4)

    wb.save(cfg.out_pend)


# ===========================================================================
# 7. Salida 3: Resumen de conciliacion de SALDOS (5 hojas, formato clasico)
# ===========================================================================
def generar_resumen_saldos(conta, banco, conta_match, banco_match, grupos, cfg):
    if cfg.saldo_contable is None:
        print("  (sin --saldo-contable: se omite el Resumen de saldos)")
        return
    gb = set()
    for js in grupos.values():
        gb.update(js)
    conta_pend = [c for i, c in enumerate(conta)
                  if not c.get("anulado") and not conta_match[i]]
    banco_pend = [b for j, b in enumerate(banco)
                  if not b.get("anulado") and not banco_match[j] and j not in gb]

    dep = sorted([c for c in conta_pend if c["tipo_bco"] == "C"], key=lambda x: -x["importe"])
    cheq = sorted([c for c in conta_pend if c["tipo_bco"] == "D"], key=lambda x: -x["importe"])
    cred = sorted([b for b in banco_pend if b["tipo"] == "C"], key=lambda x: -x["importe"])
    deb = sorted([b for b in banco_pend if b["tipo"] == "D"], key=lambda x: -x["importe"])

    t_dep = sum(c["importe"] for c in dep)
    t_cheq = sum(c["importe"] for c in cheq)
    t_cred = sum(b["importe"] for b in cred)
    t_deb = sum(b["importe"] for b in deb)
    saldo_calc = cfg.saldo_banco + t_dep - t_cheq - t_cred + t_deb
    dif = saldo_calc - cfg.saldo_contable

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conciliacion %s" % cfg.mmyyyy
    ws.append(["CONCILIACION BANCARIA - BANCO CREDICOOP - %s %d" % (cfg.month_name, cfg.year), ""])
    ws.append(["", ""])
    filas = [
        ("Saldo segun extracto bancario %s" % cfg.saldo_label, cfg.saldo_banco, None),
        ("(+) Depositos en transito (en libros, no acreditados)", t_dep, VERDE),
        ("(-) Cheques / transf. librados no debitados", -t_cheq, None),
        ("(-) Creditos del banco no contabilizados", -t_cred, None),
        ("(+) Debitos / gastos del banco no contabilizados", t_deb, None),
        ("(=) Saldo segun contabilidad (calculado)", saldo_calc, GRIS),
        ("", None, None),
        ("Saldo contable informado", cfg.saldo_contable, None),
        ("DIFERENCIA", dif, None),
    ]
    for txt, val, fill in filas:
        ws.append([txt, val])
        r = ws.max_row
        if val is not None:
            ws.cell(r, 2).number_format = MONEY
        if fill:
            for c in (1, 2):
                ws.cell(r, c).fill = fill
        if txt.startswith("(=)") or txt == "DIFERENCIA":
            ws.cell(r, 1).font = Font(bold=True)
            ws.cell(r, 2).font = Font(bold=True)
    ws.cell(1, 1).font = Font(bold=True, size=13, color="1A1F5E")

    nota = ("NOTA: la conciliacion parte del saldo del extracto y aplica las partidas "
            "no conciliadas del mes. Si la DIFERENCIA no es 0, corresponde al arrastre "
            "de partidas conciliatorias de meses anteriores y/o operaciones pendientes "
            "de revisar. Los (+) Depositos en transito y (-) Creditos del banco no "
            "contabilizados suelen ser los MISMOS depositos diferidos (VTC en libros vs "
            "'Gestion de Documentos Diferidos' en el banco) que no aparean 1:1 y se compensan.")
    ws.append(["", ""])
    ws.append([nota, ""])
    ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
    ws.row_dimensions[ws.max_row].height = 90
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 22

    def hoja_conta(titulo, items):
        w = wb.create_sheet(titulo)
        w.append(["Fecha", "Clase", "Numero", "Importe", "Sujeto"])
        _style_header(w, 5)
        for c in items:
            w.append([c["fecha"], c["clase"], c["numero"], c["importe"], c["sujeto"]])
            w.cell(w.max_row, 1).number_format = "DD/MM/YYYY"
            w.cell(w.max_row, 4).number_format = MONEY
            for cc in range(1, 6):
                w.cell(w.max_row, cc).border = THIN
        w.append(["", "", "TOTAL", sum(c["importe"] for c in items), ""])
        for cc in range(1, 6):
            w.cell(w.max_row, cc).font = Font(bold=True)
            w.cell(w.max_row, cc).fill = AMAR
        w.cell(w.max_row, 4).number_format = MONEY
        _autofit(w)

    def hoja_banco(titulo, items, con_categoria=False):
        cols = ["Fecha", "Combte", "Descripcion", "Importe"]
        if con_categoria:
            cols.append("Categoria")
        w = wb.create_sheet(titulo)
        w.append(cols)
        _style_header(w, len(cols))
        for b in items:
            row = [b["fecha"], b["combte"], b["desc"], b["importe"]]
            if con_categoria:
                row.append(categoria_banco(b["desc"]))
            w.append(row)
            w.cell(w.max_row, 1).number_format = "DD/MM/YYYY"
            w.cell(w.max_row, 4).number_format = MONEY
            for cc in range(1, len(cols) + 1):
                w.cell(w.max_row, cc).border = THIN
        w.append(["", "", "TOTAL", sum(b["importe"] for b in items)] + ([""] if con_categoria else []))
        for cc in range(1, len(cols) + 1):
            w.cell(w.max_row, cc).font = Font(bold=True)
            w.cell(w.max_row, cc).fill = AMAR
        w.cell(w.max_row, 4).number_format = MONEY
        if con_categoria:
            sub = defaultdict(float)
            for b in items:
                sub[categoria_banco(b["desc"])] += b["importe"]
            w.append(["", "", "", "", ""])
            for cat, imp in sorted(sub.items(), key=lambda x: -x[1]):
                w.append(["", "Subtotal:", cat, imp, ""])
                w.cell(w.max_row, 4).number_format = MONEY
                w.cell(w.max_row, 3).font = Font(italic=True)
        _autofit(w)

    hoja_conta("Depositos en transito", dep)
    hoja_conta("Cheques no debitados", cheq)
    hoja_banco("Creditos banco no contab", cred)
    hoja_banco("Gastos-debitos a registrar", deb, con_categoria=True)

    wb.save(cfg.out_resumen)
    print("  Saldo banco %s: %.2f | calc contable: %.2f | informado: %.2f | DIF: %.2f"
          % (cfg.saldo_label, cfg.saldo_banco, saldo_calc, cfg.saldo_contable, dif))


# ===========================================================================
def run(cfg):
    print("Periodo:", cfg.mmyyyy, "| corte:", cfg.stop_marker)
    print("Conta   :", cfg.conta_xlsx)
    print("Extracto:", cfg.extracto_pdf)

    banco = parse_extracto(cfg.extracto_pdf, cfg.stop_marker)
    conta = parse_contabilidad(cfg.conta_xlsx)
    print("  banco:", len(banco), "| conta:", len(conta))

    if cfg.saldo_banco is None:
        cfg.saldo_banco = saldo_final_pdf(cfg.extracto_pdf, cfg.stop_marker)
        if cfg.saldo_banco is None:
            print("  AVISO: no pude leer el saldo bancario del PDF (marcador %s)" % cfg.stop_marker)

    conta_anul = marcar_anulados(conta, lambda r: "IN" if r["debe"] > 0 else "OUT")
    banco_anul = marcar_anulados(banco, lambda r: "IN" if r["tipo"] == "C" else "OUT")

    cm, bm, _ = conciliar(conta, banco)
    grupos = conciliar_grupos(conta, banco, cm, bm)

    generar_excel(conta, banco, cm, bm, cfg)
    print("OK ->", cfg.out_conc)
    generar_excel_pendientes(conta, banco, cm, bm, grupos, cfg, conta_anul, banco_anul)
    print("OK ->", cfg.out_pend)
    if cfg.saldo_contable is not None and cfg.saldo_banco is not None:
        generar_resumen_saldos(conta, banco, cm, bm, grupos, cfg)
        print("OK ->", cfg.out_resumen)


def main():
    ap = argparse.ArgumentParser(description="MVP Conciliacion Bancaria Credicoop")
    ap.add_argument("--periodo", required=True, help="YYYY-MM (ej. 2025-07)")
    ap.add_argument("--saldo-contable", type=float, default=None,
                    help="Saldo contable de cierre (libro banco). Si falta, no genera el Resumen de saldos")
    ap.add_argument("--saldo-banco", type=float, default=None,
                    help="Saldo bancario de cierre (opcional; por defecto se lee del PDF)")
    ap.add_argument("--base", default=BASE_DIR_DEFAULT, help="Carpeta base de datos")
    ap.add_argument("--conta", default=None, help="Ruta Excel de contabilidad (opcional)")
    ap.add_argument("--extracto", default=None, help="Ruta PDF de extracto (opcional)")
    a = ap.parse_args()
    cfg = Config(a.periodo, a.saldo_contable, a.base,
                 conta=a.conta, extracto=a.extracto, saldo_banco=a.saldo_banco)
    run(cfg)


if __name__ == "__main__":
    main()
