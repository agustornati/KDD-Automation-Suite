"""Exportadores del módulo de conciliación bancaria.

El :class:`ExcelExporter` reconstruye los tres libros Excel a partir de un
:class:`ReconciliationResult`, con el mismo contenido y formato que el script
legacy. Está desacoplado del motor: recibe el resultado ya calculado.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from shared.exporters.base import ResultExporter

from .models import ReconciliationResult
from .utils import CATEGORIA_GASTO, categoria_banco, categoria_banco_detalle

# --- Estilos (idénticos al legacy) -----------------------------------------
VERDE = PatternFill("solid", fgColor="C6EFCE")
ROJO = PatternFill("solid", fgColor="FFC7CE")
GRIS = PatternFill("solid", fgColor="F2F2F2")
HDR = PatternFill("solid", fgColor="1A1F5E")
HDRF = Font(bold=True, color="FFFFFF")
AMAR = PatternFill("solid", fgColor="FFE600")
NARA = PatternFill("solid", fgColor="F4B942")   # naranja — SALDO-RECOVER
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)
MONEY = "#,##0.00"


def _style_header(ws, ncols: int) -> None:
    """Aplica el estilo de encabezado y congela la primera fila."""
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR
        cell.font = HDRF
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autofit(ws) -> None:
    """Ajusta el ancho de las columnas al contenido."""
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 55)


class ExcelExporter(ResultExporter):
    """Genera los tres libros Excel de la conciliación."""

    def export(self, result: ReconciliationResult, output_dir: Path) -> list[Path]:
        """Escribe los Excel en ``output_dir`` y devuelve sus rutas.

        Genera siempre la Conciliación (verde/rojo) y los No Coincidentes. El
        Resumen de saldos se genera si hay saldo bancario disponible.
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        mmyyyy = result.period.mmyyyy

        out_conc = output_dir / f"Conciliacion_Credicoop_{mmyyyy}.xlsx"
        out_pend = output_dir / f"Conciliacion_NoCoincidentes_{mmyyyy}.xlsx"
        out_resumen = output_dir / f"Resumen_Conciliacion_{mmyyyy}.xlsx"

        self._generar_conciliacion(result, out_conc)
        self._generar_no_coincidentes(result, out_pend)

        archivos = [out_conc, out_pend]
        if result.data.saldo_banco is not None:
            self._generar_resumen_saldos(result, out_resumen)
            archivos.append(out_resumen)
        return archivos

    # -- Salida 1: Conciliación (verde/rojo) + Resumen ----------------------
    def _generar_conciliacion(self, result: ReconciliationResult, path: Path) -> None:
        data = result.data
        conta, banco = data.conta, data.banco
        conta_match, banco_match = data.conta_match, data.banco_match
        period = result.period
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "Contabilidad"
        cols = ["Fecha", "Clase", "Numero", "Debe", "Haber", "Sujeto",
                "Lado banco", "Estado"]
        ws.append(cols)
        _style_header(ws, len(cols))
        for i, c in enumerate(conta):
            if c.get("anulado"):
                continue
            estado = "CONCILIADO" if conta_match[i] else "NO CONCILIADO"
            ws.append([c["fecha"], c["clase"], c["numero"], c["debe"] or None,
                       c["haber"] or None, c["sujeto"],
                       "Credito" if c["tipo_bco"] == "C" else "Debito", estado])
            fill = VERDE if conta_match[i] else ROJO
            for cc in range(1, len(cols) + 1):
                ws.cell(ws.max_row, cc).fill = fill
                ws.cell(ws.max_row, cc).border = THIN
            ws.cell(ws.max_row, 1).number_format = "DD/MM/YYYY"
            for cc in (4, 5):
                ws.cell(ws.max_row, cc).number_format = MONEY
        _autofit(ws)

        ws2 = wb.create_sheet("Extracto Banco")
        cols2 = ["Fecha", "Combte", "Descripcion", "Debito", "Credito",
                 "Estado", "Categoria"]
        ws2.append(cols2)
        _style_header(ws2, len(cols2))
        for j, b in enumerate(banco):
            if b.get("anulado"):
                continue
            estado = "CONCILIADO" if banco_match[j] else "NO CONCILIADO"
            ws2.append([b["fecha"], b["combte"], b["desc"],
                        b["importe"] if b["tipo"] == "D" else None,
                        b["importe"] if b["tipo"] == "C" else None, estado,
                        categoria_banco(b["desc"])])
            fill = VERDE if banco_match[j] else ROJO
            for cc in range(1, len(cols2) + 1):
                ws2.cell(ws2.max_row, cc).fill = fill
                ws2.cell(ws2.max_row, cc).border = THIN
            ws2.cell(ws2.max_row, 1).number_format = "DD/MM/YYYY"
            for cc in (4, 5):
                ws2.cell(ws2.max_row, cc).number_format = MONEY
        _autofit(ws2)

        ws3 = wb.create_sheet("Resumen")
        s = result.stats
        resumen = [
            ["CONCILIACION BANCARIA - BANCO CREDICOOP - %02d/%04d"
             % (period.month, period.year), ""],
            ["", ""],
            ["CONTABILIDAD", ""],
            ["  Movimientos", s.movimientos_sap],
            ["  Conciliados", s.sap_conciliados],
            ["  No conciliados", s.sap_pendientes],
            ["  Total Debe", s.total_debe],
            ["  Total Haber", s.total_haber],
            ["", ""],
            ["EXTRACTO BANCO", ""],
            ["  Movimientos", s.movimientos_banco],
            ["  Conciliados", s.banco_conciliados],
            ["  No conciliados", s.banco_pendientes],
            ["    de los cuales operaciones a revisar", s.banco_pend_operaciones],
            ["    de los cuales gastos/impuestos bancarios", s.banco_pend_gastos],
            ["  Total Debito", s.total_debito_banco],
            ["  Total Credito", s.total_credito_banco],
            ["", ""],
            ["Importe conciliado (coincidente)", s.importe_conciliado],
            ["Regla", "Conta DEBE <-> Banco CREDITO  |  Conta HABER <-> Banco DEBITO"],
        ]
        for row in resumen:
            ws3.append(row)
        ws3.cell(1, 1).font = Font(bold=True, size=13, color="1A1F5E")
        for r in (3, 10):
            ws3.cell(r, 1).font = Font(bold=True, color="2D3277")
        for r in range(1, ws3.max_row + 1):
            v = ws3.cell(r, 2).value
            if isinstance(v, float):
                ws3.cell(r, 2).number_format = MONEY
        ws3.column_dimensions["A"].width = 42
        ws3.column_dimensions["B"].width = 28

        wb.save(path)

    # -- Salida 2: No coincidentes (pendientes + gastos + anulados) ---------
    def _generar_no_coincidentes(self, result: ReconciliationResult, path: Path) -> None:
        data = result.data
        conta, banco = data.conta, data.banco
        conta_match, banco_match = data.conta_match, data.banco_match
        grupos = data.grupos
        wb = openpyxl.Workbook()
        etiqueta = {i: "%s %s" % (conta[i]["clase"], conta[i]["numero"]) for i in grupos}
        banco_a_conta: dict[int, str] = {}
        for i, js in grupos.items():
            for j in js:
                banco_a_conta[j] = etiqueta[i]

        ws = wb.active
        ws.title = "Contabilidad no 1a1"
        cols = ["Fecha", "Clase", "Numero", "Debe", "Haber", "Sujeto",
                "Lado banco", "Estado", "Detalle grupo (banco)"]
        ws.append(cols)
        _style_header(ws, len(cols))
        for i, c in enumerate(conta):
            if c.get("anulado"):
                continue
            if conta_match[i] is True:
                continue
            if conta_match[i] == "GRUPO":
                estado = "CONCILIADO POR SUMA"
                det = "  +  ".join(
                    "{} ${:,.2f}".format(banco[j]["fecha"].strftime("%d/%m"),
                                         banco[j]["importe"])
                    for j in grupos[i])
                fill = VERDE
            else:
                estado, det, fill = "NO CONCILIADO", "", ROJO
            ws.append([c["fecha"], c["clase"], c["numero"], c["debe"] or None,
                       c["haber"] or None, c["sujeto"],
                       "Credito" if c["tipo_bco"] == "C" else "Debito", estado, det])
            for cc in range(1, len(cols) + 1):
                ws.cell(ws.max_row, cc).fill = fill
                ws.cell(ws.max_row, cc).border = THIN
            ws.cell(ws.max_row, 1).number_format = "DD/MM/YYYY"
            for cc in (4, 5):
                ws.cell(ws.max_row, cc).number_format = MONEY
        _autofit(ws)

        ws2 = wb.create_sheet("Extracto no 1a1")
        cols2 = ["Fecha", "Combte", "Descripcion", "Debito", "Credito",
                 "Categoria", "Estado", "Grupo (conta)"]
        ws2.append(cols2)
        _style_header(ws2, len(cols2))
        for j, b in enumerate(banco):
            if b.get("anulado"):
                continue
            if banco_match[j] is True and j not in banco_a_conta:
                continue
            if j not in banco_a_conta and categoria_banco(b["desc"]) == CATEGORIA_GASTO:
                continue
            if j in banco_a_conta:
                estado, grp, fill = "CONCILIADO POR SUMA", banco_a_conta[j], VERDE
            else:
                estado, grp, fill = "NO CONCILIADO", "", ROJO
            ws2.append([b["fecha"], b["combte"], b["desc"],
                        b["importe"] if b["tipo"] == "D" else None,
                        b["importe"] if b["tipo"] == "C" else None,
                        categoria_banco(b["desc"]), estado, grp])
            for cc in range(1, len(cols2) + 1):
                ws2.cell(ws2.max_row, cc).fill = fill
                ws2.cell(ws2.max_row, cc).border = THIN
            ws2.cell(ws2.max_row, 1).number_format = "DD/MM/YYYY"
            for cc in (4, 5):
                ws2.cell(ws2.max_row, cc).number_format = MONEY
        _autofit(ws2)

        ws3 = wb.create_sheet("Gastos a registrar")
        resumen: dict = defaultdict(lambda: [0, 0.0])
        for j, b in enumerate(banco):
            if banco_match[j] is True and j not in banco_a_conta:
                continue
            if j in banco_a_conta:
                continue
            label = categoria_banco_detalle(b["desc"])
            if label is None:
                continue
            key = label if label != CATEGORIA_GASTO else (b["desc"] or "(sin descripcion)")
            resumen[key][0] += 1
            resumen[key][1] += b["importe"]
        cols3 = ["Concepto (descripcion)", "Cant. movimientos", "Importe total"]
        ws3.append(cols3)
        _style_header(ws3, len(cols3))
        total = 0.0
        for desc, (cant, imp) in sorted(resumen.items(), key=lambda x: -x[1][1]):
            ws3.append([desc, cant, imp])
            total += imp
            for cc in range(1, len(cols3) + 1):
                ws3.cell(ws3.max_row, cc).border = THIN
            ws3.cell(ws3.max_row, 3).number_format = MONEY
        ws3.append(["TOTAL", sum(c for c, _ in resumen.values()), total])
        for cc in range(1, len(cols3) + 1):
            ws3.cell(ws3.max_row, cc).font = Font(bold=True)
            ws3.cell(ws3.max_row, cc).fill = AMAR
        ws3.cell(ws3.max_row, 3).number_format = MONEY
        _autofit(ws3)

        ws4 = wb.create_sheet("Anulados (salio y volvio)")
        cols4 = ["Fuente", "Fecha", "Detalle", "Importe", "Contrapartida"]
        ws4.append(cols4)
        _style_header(ws4, len(cols4))
        for a, b in data.conta_anul:
            ws4.append(["Contabilidad", conta[a]["fecha"],
                        "%s %s (Haber) / %s %s (Debe)" % (
                            conta[a]["clase"], conta[a]["numero"],
                            conta[b]["clase"], conta[b]["numero"]),
                        conta[a]["importe"], "Netea contra si mismo"])
        for a, b in data.banco_anul:
            ws4.append(["Extracto banco", banco[a]["fecha"],
                        "%s (Debito) / %s (Credito)" % (banco[a]["desc"],
                                                        banco[b]["desc"]),
                        banco[a]["importe"], "Cheque depositado y rechazado"])
        for r in range(2, ws4.max_row + 1):
            ws4.cell(r, 4).number_format = MONEY
            for cc in range(1, len(cols4) + 1):
                ws4.cell(r, cc).border = THIN
            ws4.cell(r, 2).number_format = "DD/MM/YYYY"
        _autofit(ws4)

        wb.save(path)

    # -- Salida 3: Resumen de conciliación de SALDOS (5 hojas) --------------
    def _generar_resumen_saldos(self, result: ReconciliationResult, path: Path) -> None:
        data = result.data
        conta, banco = data.conta, data.banco
        conta_match, banco_match = data.conta_match, data.banco_match
        grupos = data.grupos
        period = result.period
        saldo_banco = data.saldo_banco
        saldo_contable = data.saldo_contable

        gb: set = set()
        for js in grupos.values():
            gb.update(js)
        conta_pend = [c for i, c in enumerate(conta)
                      if not c.get("anulado") and not conta_match[i]]
        banco_pend = [b for j, b in enumerate(banco)
                      if not b.get("anulado") and not banco_match[j] and j not in gb]

        dep = sorted([c for c in conta_pend if c["tipo_bco"] == "C"], key=lambda x: -x["importe"])
        cheq = sorted([c for c in conta_pend if c["tipo_bco"] == "D"], key=lambda x: -x["importe"])
        cred = sorted([b for b in banco_pend if b["tipo"] == "C"], key=lambda x: -x["importe"])
        deb = sorted([b for b in banco_pend if b["tipo"] == "D"], key=lambda x: -x["importe"])

        t_dep = sum(c["importe"] for c in dep)
        t_cheq = sum(c["importe"] for c in cheq)
        t_cred = sum(b["importe"] for b in cred)
        t_deb = sum(b["importe"] for b in deb)
        saldo_calc = saldo_banco + t_dep - t_cheq - t_cred + t_deb
        dif = saldo_calc - saldo_contable

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Conciliacion %s" % period.mmyyyy
        ws.append(["CONCILIACION BANCARIA - BANCO CREDICOOP - %s %d"
                   % (period.month_name, period.year), ""])
        ws.append(["", ""])
        filas = [
            ("Saldo segun extracto bancario %s" % period.saldo_label, saldo_banco, None),
            ("(+) Depositos en transito (en libros, no acreditados)", t_dep, VERDE),
            ("(-) Cheques / transf. librados no debitados", -t_cheq, None),
            ("(-) Creditos del banco no contabilizados", -t_cred, None),
            ("(+) Debitos / gastos del banco no contabilizados", t_deb, None),
            ("(=) Saldo segun contabilidad (calculado)", saldo_calc, GRIS),
            ("", None, None),
            ("Saldo contable informado", saldo_contable, None),
            ("DIFERENCIA", dif, None),
        ]
        for txt, val, fill in filas:
            ws.append([txt, val])
            r = ws.max_row
            if val is not None:
                ws.cell(r, 2).number_format = MONEY
            if fill:
                for c in (1, 2):
                    ws.cell(r, c).fill = fill
            if txt.startswith("(=)") or txt == "DIFERENCIA":
                ws.cell(r, 1).font = Font(bold=True)
                ws.cell(r, 2).font = Font(bold=True)
        ws.cell(1, 1).font = Font(bold=True, size=13, color="1A1F5E")

        nota = ("NOTA: la conciliacion parte del saldo del extracto y aplica las partidas "
                "no conciliadas del mes. Si la DIFERENCIA no es 0, corresponde al arrastre "
                "de partidas conciliatorias de meses anteriores y/o operaciones pendientes "
                "de revisar. Los (+) Depositos en transito y (-) Creditos del banco no "
                "contabilizados suelen ser los MISMOS depositos diferidos (VTC en libros vs "
                "'Gestion de Documentos Diferidos' en el banco) que no aparean 1:1 y se compensan.")
        ws.append(["", ""])
        ws.append([nota, ""])
        ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        ws.row_dimensions[ws.max_row].height = 90
        ws.column_dimensions["A"].width = 55
        ws.column_dimensions["B"].width = 22

        def hoja_conta(titulo, items):
            w = wb.create_sheet(titulo)
            w.append(["Fecha", "Clase", "Numero", "Importe", "Sujeto"])
            _style_header(w, 5)
            for c in items:
                w.append([c["fecha"], c["clase"], c["numero"], c["importe"], c["sujeto"]])
                w.cell(w.max_row, 1).number_format = "DD/MM/YYYY"
                w.cell(w.max_row, 4).number_format = MONEY
                for cc in range(1, 6):
                    w.cell(w.max_row, cc).border = THIN
            w.append(["", "", "TOTAL", sum(c["importe"] for c in items), ""])
            for cc in range(1, 6):
                w.cell(w.max_row, cc).font = Font(bold=True)
                w.cell(w.max_row, cc).fill = AMAR
            w.cell(w.max_row, 4).number_format = MONEY
            _autofit(w)

        def hoja_banco(titulo, items):
            cols = ["Fecha", "Combte", "Descripcion", "Importe", "Nota"]
            w = wb.create_sheet(titulo)
            w.append(cols)
            _style_header(w, len(cols))
            for b in items:
                desc = b["desc"] or ""
                is_recover = "SALDO-RECOVER" in desc
                if is_recover:
                    tipo = "CREDITO" if "CRED" in desc else "DEBITO"
                    nota = ("VERIFICAR MANUALMENTE: transaccion no leida por OCR. "
                            "Buscar en el extracto impreso el %s de ~$%s" % (
                                tipo,
                                "{:,.2f}".format(b["importe"])))
                else:
                    nota = ""
                w.append([b["fecha"], b["combte"], desc, b["importe"], nota])
                w.cell(w.max_row, 1).number_format = "DD/MM/YYYY"
                w.cell(w.max_row, 4).number_format = MONEY
                fill = NARA if is_recover else None
                for cc in range(1, len(cols) + 1):
                    w.cell(w.max_row, cc).border = THIN
                    if fill:
                        w.cell(w.max_row, cc).fill = fill
                if nota:
                    w.cell(w.max_row, 5).alignment = Alignment(wrap_text=False)
            w.append(["", "", "TOTAL", sum(b["importe"] for b in items), ""])
            for cc in range(1, len(cols) + 1):
                w.cell(w.max_row, cc).font = Font(bold=True)
                w.cell(w.max_row, cc).fill = AMAR
            w.cell(w.max_row, 4).number_format = MONEY
            _autofit(w)

        def hoja_gastos_agrupados(titulo, items):
            """Agrupa débitos pendientes por nombre canónico de gasto."""
            resumen_g: dict[str, list] = {}
            for b in items:
                label = categoria_banco_detalle(b["desc"])
                if label is None:
                    continue
                key = label if label != CATEGORIA_GASTO else (b["desc"] or "(sin descripcion)")
                acc = resumen_g.setdefault(key, [0, 0.0])
                acc[0] += 1
                acc[1] += b["importe"]
            cols = ["Concepto", "Cant. movimientos", "Importe total"]
            w = wb.create_sheet(titulo)
            w.append(cols)
            _style_header(w, len(cols))
            total = 0.0
            for label, (cant, imp) in sorted(resumen_g.items(), key=lambda x: -x[1][1]):
                w.append([label, cant, imp])
                total += imp
                w.cell(w.max_row, 3).number_format = MONEY
                for cc in range(1, len(cols) + 1):
                    w.cell(w.max_row, cc).border = THIN
            w.append(["TOTAL", sum(v[0] for v in resumen_g.values()), total])
            for cc in range(1, len(cols) + 1):
                w.cell(w.max_row, cc).font = Font(bold=True)
                w.cell(w.max_row, cc).fill = AMAR
            w.cell(w.max_row, 3).number_format = MONEY
            _autofit(w)

        deb_oper   = [b for b in deb if categoria_banco_detalle(b["desc"]) is None]
        deb_gastos = [b for b in deb if categoria_banco_detalle(b["desc"]) is not None]

        hoja_conta("Depositos en transito", dep)
        hoja_conta("Cheques no debitados", cheq)
        hoja_banco("Creditos banco no contab", cred)
        hoja_banco("Operaciones banco no contab", deb_oper)
        hoja_gastos_agrupados("Gastos-debitos a registrar", deb_gastos)

        # Hoja resumen de entradas SALDO-RECOVER para verificacion manual
        todos_banco = list(banco)
        recovers = sorted(
            [b for b in todos_banco if "SALDO-RECOVER" in (b["desc"] or "")],
            key=lambda b: -b["importe"],
        )
        if recovers:
            wr = wb.create_sheet("SALDO-RECOVER (verificar)")
            cols_r = ["Fecha", "Tipo", "Importe", "Nota para la contadora"]
            wr.append(cols_r)
            _style_header(wr, len(cols_r))
            for b in recovers:
                desc = b["desc"] or ""
                tipo = "CREDITO banco" if "CRED" in desc else "DEBITO banco"
                nota = ("Transaccion no leida por OCR. "
                        "Verificar en el extracto impreso la fecha %s y "
                        "ubicar el %s de ~$%s" % (
                            b["fecha"].strftime("%d/%m") if b["fecha"] else "?",
                            tipo,
                            "{:,.2f}".format(b["importe"])))
                wr.append([b["fecha"], tipo, b["importe"], nota])
                wr.cell(wr.max_row, 1).number_format = "DD/MM/YYYY"
                wr.cell(wr.max_row, 3).number_format = MONEY
                for cc in range(1, len(cols_r) + 1):
                    wr.cell(wr.max_row, cc).border = THIN
                    wr.cell(wr.max_row, cc).fill = NARA
            wr.append(["", "TOTAL", sum(b["importe"] for b in recovers), ""])
            for cc in range(1, len(cols_r) + 1):
                wr.cell(wr.max_row, cc).font = Font(bold=True)
                wr.cell(wr.max_row, cc).fill = AMAR
            wr.cell(wr.max_row, 3).number_format = MONEY
            wr.column_dimensions["A"].width = 12
            wr.column_dimensions["B"].width = 16
            wr.column_dimensions["C"].width = 18
            wr.column_dimensions["D"].width = 80
            wr.cell(1, 4).alignment = Alignment(wrap_text=False)

        wb.save(path)
